"""
產生金酒成本彙整系統的虛擬資料
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')

PRODUCTS = [
    {'code': 'KL-58-05', 'name': '金門高粱酒 58度 0.5L', 'unit': '瓶'},
    {'code': 'KL-58-075', 'name': '金門高粱酒 58度 0.75L', 'unit': '瓶'},
    {'code': 'KL-38-05', 'name': '金門高粱酒 38度 0.5L', 'unit': '瓶'},
    {'code': 'KL-52-06', 'name': '金門特級高粱酒 52度 0.6L', 'unit': '瓶'},
]

DEPARTMENTS = ['金城廠釀造部', '金城廠包裝部', '金寧廠釀造部', '金寧廠包裝部']

EMPLOYEES = [
    {'id': 'E001', 'name': '王大明', 'dept': '金城廠釀造部', 'title': '作業員'},
    {'id': 'E002', 'name': '李小花', 'dept': '金城廠釀造部', 'title': '作業員'},
    {'id': 'E003', 'name': '張志豪', 'dept': '金城廠釀造部', 'title': '領班'},
    {'id': 'E004', 'name': '陳美玲', 'dept': '金城廠包裝部', 'title': '作業員'},
    {'id': 'E005', 'name': '林建國', 'dept': '金城廠包裝部', 'title': '作業員'},
    {'id': 'E006', 'name': '黃雅慧', 'dept': '金城廠包裝部', 'title': '領班'},
    {'id': 'E007', 'name': '吳正義', 'dept': '金寧廠釀造部', 'title': '作業員'},
    {'id': 'E008', 'name': '蔡淑芬', 'dept': '金寧廠釀造部', 'title': '作業員'},
    {'id': 'E009', 'name': '鄭文傑', 'dept': '金寧廠釀造部', 'title': '領班'},
    {'id': 'E010', 'name': '許美雪', 'dept': '金寧廠包裝部', 'title': '作業員'},
    {'id': 'E011', 'name': '洪志明', 'dept': '金寧廠包裝部', 'title': '作業員'},
    {'id': 'E012', 'name': '謝佳宜', 'dept': '金寧廠包裝部', 'title': '領班'},
]

MONTH = '2026年05月'


def style_header(ws, row, col_count, title, fill_color='1F4E79'):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)


def style_col_headers(ws, row, headers, fill_color='BDD7EE'):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell, number_format=None):
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal='center')
    if number_format:
        cell.number_format = number_format


# ── 1. 薪資統計表 ──────────────────────────────────────────────
def gen_salary():
    wb = Workbook()
    ws = wb.active
    ws.title = '薪資統計表'
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    style_header(ws, 1, 8, f'金酒公司薪資統計表　{MONTH}')
    headers = ['員工編號', '姓名', '部門', '職稱', '本薪', '加班費', '其他津貼', '合計薪資']
    style_col_headers(ws, 2, headers)

    salary_data = [
        ('E001', '王大明', '金城廠釀造部', '作業員', 38000, 3200, 1000),
        ('E002', '李小花', '金城廠釀造部', '作業員', 36000, 2800, 1000),
        ('E003', '張志豪', '金城廠釀造部', '領班',   48000, 4500, 2000),
        ('E004', '陳美玲', '金城廠包裝部', '作業員', 36000, 2500, 1000),
        ('E005', '林建國', '金城廠包裝部', '作業員', 37000, 3100, 1000),
        ('E006', '黃雅慧', '金城廠包裝部', '領班',   47000, 4200, 2000),
        ('E007', '吳正義', '金寧廠釀造部', '作業員', 38000, 3500, 1000),
        ('E008', '蔡淑芬', '金寧廠釀造部', '作業員', 36500, 2900, 1000),
        ('E009', '鄭文傑', '金寧廠釀造部', '領班',   49000, 5000, 2000),
        ('E010', '許美雪', '金寧廠包裝部', '作業員', 36000, 2600, 1000),
        ('E011', '洪志明', '金寧廠包裝部', '作業員', 37500, 3000, 1000),
        ('E012', '謝佳宜', '金寧廠包裝部', '領班',   48000, 4800, 2000),
    ]

    for i, (eid, name, dept, title, base, ot, other) in enumerate(salary_data, 3):
        total = base + ot + other
        row_data = [eid, name, dept, title, base, ot, other, total]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            style_data_cell(cell)
            if j >= 5:
                cell.number_format = '#,##0'
        alt = i % 2 == 0
        if alt:
            for j in range(1, 9):
                ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    # 合計列
    total_row = len(salary_data) + 3
    ws.cell(row=total_row, column=1, value='合計').font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    for col in range(5, 9):
        col_letter = chr(64 + col)
        cell = ws.cell(row=total_row, column=col,
                       value=f'=SUM({col_letter}3:{col_letter}{total_row-1})')
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'
        cell.fill = PatternFill(fill_type='solid', fgColor='D6E4F0')
        style_data_cell(cell)

    path = os.path.join(OUTPUT_DIR, '01_薪資統計表.xlsx')
    wb.save(path)
    print(f'✓ {path}')


# ── 2. 工時比例表 ──────────────────────────────────────────────
def gen_workhour():
    wb = Workbook()
    ws = wb.active
    ws.title = '工時比例表'
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16

    style_header(ws, 1, 6, f'金酒公司直工工時比例表　{MONTH}')
    headers = ['部門', '總出勤工時', 'KL-58-05\n工時分配', 'KL-58-075\n工時分配',
               'KL-38-05\n工時分配', 'KL-52-06\n工時分配']
    style_col_headers(ws, 2, headers)

    workhour_data = [
        ('金城廠釀造部', 1840, 530, 420, 580, 310),
        ('金城廠包裝部', 1760, 480, 390, 560, 330),
        ('金寧廠釀造部', 1920, 510, 450, 590, 370),
        ('金寧廠包裝部', 1800, 490, 410, 570, 330),
    ]

    for i, row_data in enumerate(workhour_data, 3):
        dept, total, *hours = row_data
        ws.cell(row=i, column=1, value=dept)
        ws.cell(row=i, column=2, value=total)
        for j, h in enumerate(hours, 3):
            ws.cell(row=i, column=j, value=h)
        for col in range(1, 7):
            style_data_cell(ws.cell(row=i, column=col), '#,##0' if col >= 2 else None)
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    path = os.path.join(OUTPUT_DIR, '02_工時比例表.xlsx')
    wb.save(path)
    print(f'✓ {path}')


# ── 3. 半成品盤存表 ──────────────────────────────────────────────
def gen_wip_stock():
    wb = Workbook()
    ws = wb.active
    ws.title = '半成品盤存表'
    for col, w in zip(['A','B','C','D','E','F','G'], [12,20,10,12,12,12,12]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 7, f'金酒公司半成品盤存表　{MONTH}')
    headers = ['品項代碼', '品項名稱', '單位', '期初存量\n(公升)', '本期產量\n(公升)',
               '本期移轉\n(公升)', '期末存量\n(公升)']
    style_col_headers(ws, 2, headers)

    wip_data = [
        ('WIP-58', '高粱酒基酒 58度', '公升', 12500, 45000, 43200, 14300),
        ('WIP-38', '高粱酒基酒 38度', '公升',  8200, 32000, 30500,  9700),
        ('WIP-52', '特級高粱基酒 52度', '公升', 5100, 18000, 17200,  5900),
    ]

    for i, row_data in enumerate(wip_data, 3):
        code, name, unit, open_qty, prod, transfer, close = row_data
        row = [code, name, unit, open_qty, prod, transfer, close]
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            style_data_cell(cell)
            if j >= 4:
                cell.number_format = '#,##0'
        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    path = os.path.join(OUTPUT_DIR, '03_半成品盤存表.xlsx')
    wb.save(path)
    print(f'✓ {path}')


# ── 4. 包裝廠生產月報 ──────────────────────────────────────────────
def gen_packaging():
    wb = Workbook()
    ws = wb.active
    ws.title = '包裝廠生產月報'
    for col, w in zip(['A','B','C','D','E','F','G','H'], [12,22,10,12,12,14,14,14]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 8, f'金酒公司包裝廠生產月報　{MONTH}')
    headers = ['產品代碼', '產品名稱', '單位', '計畫產量', '實際產量',
               '包裝工時\n(小時)', '瓶/箱數', '損耗數量']
    style_col_headers(ws, 2, headers)

    pkg_data = [
        ('KL-58-05',  '金門高粱酒 58度 0.5L',   '瓶', 85000,  83500, 1680, 6958, 420),
        ('KL-58-075', '金門高粱酒 58度 0.75L',   '瓶', 60000,  58800, 1560, 4900, 295),
        ('KL-38-05',  '金門高粱酒 38度 0.5L',    '瓶', 95000,  93200, 1820, 7767, 465),
        ('KL-52-06',  '金門特級高粱酒 52度 0.6L', '瓶', 42000,  41100,  980, 3425, 205),
    ]

    for i, row_data in enumerate(pkg_data, 3):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            style_data_cell(cell)
            if j >= 4:
                cell.number_format = '#,##0'
        if i % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    path = os.path.join(OUTPUT_DIR, '04_包裝廠生產月報.xlsx')
    wb.save(path)
    print(f'✓ {path}')


# ── 5. 物料盤存表 ──────────────────────────────────────────────
def gen_material():
    wb = Workbook()
    ws = wb.active
    ws.title = '物料盤存表'
    for col, w in zip(['A','B','C','D','E','F','G','H'], [12,22,8,12,12,12,12,14]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 8, f'金酒公司物料盤存表　{MONTH}')
    headers = ['物料代碼', '物料名稱', '單位', '期初庫存', '本期進貨', '本期領用', '期末庫存', '單位成本(元)']
    style_col_headers(ws, 2, headers)

    mat_data = [
        ('M001', '高粱',      '公斤', 35000, 120000, 118500, 36500,  18.5),
        ('M002', '小麥（麴用）', '公斤',  8000,  25000,  24200,  8800,  12.0),
        ('M003', '0.5L玻璃瓶', '個',  15000, 100000,  96800, 18200,   8.2),
        ('M004', '0.75L玻璃瓶', '個', 10000,  65000,  62300, 12700,  11.5),
        ('M005', '0.6L玻璃瓶', '個',   8000,  45000,  43500,  9500,  10.8),
        ('M006', '瓶蓋',       '個',  20000, 200000, 196800, 23200,   0.8),
        ('M007', '標籤紙',     '張',  30000, 250000, 243000, 37000,   0.5),
        ('M008', '外箱紙板',   '個',   5000,  25000,  24200,  5800,   9.5),
    ]

    for i, row_data in enumerate(mat_data, 3):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            style_data_cell(cell)
            if j >= 4:
                cell.number_format = '#,##0' if j < 8 else '#,##0.0'
        if i % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    path = os.path.join(OUTPUT_DIR, '05_物料盤存表.xlsx')
    wb.save(path)
    print(f'✓ {path}')


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f'\n產生虛擬資料到 {OUTPUT_DIR}/\n')
    gen_salary()
    gen_workhour()
    gen_wip_stock()
    gen_packaging()
    gen_material()
    print('\n✅ 所有虛擬資料產生完成！')
