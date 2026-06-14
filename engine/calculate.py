"""
金酒成本彙整計算引擎
讀取 5 張原始 Excel → 計算直工製造費用分攤 + 物料成本 → 輸出產品成本表
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'output')

PRODUCT_CODES = ['KL-58-05', 'KL-58-075', 'KL-38-05', 'KL-52-06']
PRODUCT_NAMES = {
    'KL-58-05':  '金門高粱酒 58度 0.5L',
    'KL-58-075': '金門高粱酒 58度 0.75L',
    'KL-38-05':  '金門高粱酒 38度 0.5L',
    'KL-52-06':  '金門特級高粱酒 52度 0.6L',
}

MATERIAL_USAGE = {
    'KL-58-05':  {'M001': 0.35, 'M002': 0.05, 'M003': 1.0, 'M006': 1.0, 'M007': 1.0, 'M008': 0.083},
    'KL-58-075': {'M001': 0.52, 'M002': 0.08, 'M004': 1.0, 'M006': 1.0, 'M007': 1.0, 'M008': 0.083},
    'KL-38-05':  {'M001': 0.22, 'M002': 0.04, 'M003': 1.0, 'M006': 1.0, 'M007': 1.0, 'M008': 0.083},
    'KL-52-06':  {'M001': 0.42, 'M002': 0.07, 'M005': 1.0, 'M006': 1.0, 'M007': 1.0, 'M008': 0.083},
}


def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)


def read_salary(path):
    df = pd.read_excel(path, header=1)
    df.columns = ['員工編號', '姓名', '部門', '職稱', '本薪', '加班費', '其他津貼', '合計薪資']
    df = df[df['員工編號'].notna() & (df['員工編號'] != '合計')]
    df['合計薪資'] = pd.to_numeric(df['合計薪資'], errors='coerce').fillna(0)
    return df


def read_workhour(path):
    df = pd.read_excel(path, header=1)
    df.columns = ['部門', '總出勤工時', 'KL-58-05', 'KL-58-075', 'KL-38-05', 'KL-52-06']
    df = df[df['部門'].notna()]
    return df


def read_packaging(path):
    df = pd.read_excel(path, header=1)
    df.columns = ['產品代碼', '產品名稱', '單位', '計畫產量', '實際產量',
                  '包裝工時', '瓶箱數', '損耗數量']
    df = df[df['產品代碼'].notna()]
    df['實際產量'] = pd.to_numeric(df['實際產量'], errors='coerce').fillna(0)
    return df


def read_material(path):
    df = pd.read_excel(path, header=1)
    df.columns = ['物料代碼', '物料名稱', '單位', '期初庫存', '本期進貨', '本期領用', '期末庫存', '單位成本']
    df = df[df['物料代碼'].notna()]
    df['單位成本'] = pd.to_numeric(df['單位成本'], errors='coerce').fillna(0)
    return df


def calculate(data_dir=None, output_dir=None):
    data_dir = data_dir or DATA_DIR
    output_dir = output_dir or OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    logs = []

    # ── 讀取原始資料 ──
    salary_df = read_salary(os.path.join(data_dir, '01_薪資統計表.xlsx'))
    workhour_df = read_workhour(os.path.join(data_dir, '02_工時比例表.xlsx'))
    packaging_df = read_packaging(os.path.join(data_dir, '04_包裝廠生產月報.xlsx'))
    material_df = read_material(os.path.join(data_dir, '05_物料盤存表.xlsx'))

    # ── 計算直工薪資分攤 ──
    dept_salary = salary_df.groupby('部門')['合計薪資'].sum().to_dict()

    labor_cost = {code: 0.0 for code in PRODUCT_CODES}
    labor_detail = []

    for _, row in workhour_df.iterrows():
        dept = row['部門']
        total_hr = row['總出勤工時']
        salary = dept_salary.get(dept, 0)
        if total_hr == 0:
            continue
        rate_per_hr = salary / total_hr

        for code in PRODUCT_CODES:
            hrs = row.get(code, 0)
            cost = rate_per_hr * hrs
            labor_cost[code] += cost
            labor_detail.append({
                '部門': dept,
                '產品代碼': code,
                '分配工時(時)': hrs,
                '薪資分攤(元)': round(cost, 0),
            })

    logs.append({'step': '直工薪資分攤', 'status': '完成',
                 'detail': f'總薪資 {sum(dept_salary.values()):,.0f} 元'})

    # ── 計算物料成本（每瓶） ──
    mat_cost_map = material_df.set_index('物料代碼')['單位成本'].to_dict()

    prod_vol = packaging_df.set_index('產品代碼')['實際產量'].to_dict()

    material_cost = {code: 0.0 for code in PRODUCT_CODES}
    material_detail = []

    for code in PRODUCT_CODES:
        qty = prod_vol.get(code, 0)
        if qty == 0:
            continue
        usage = MATERIAL_USAGE.get(code, {})
        for mat_code, usage_per_unit in usage.items():
            unit_cost = mat_cost_map.get(mat_code, 0)
            total_mat_cost = unit_cost * usage_per_unit * qty
            material_cost[code] += total_mat_cost
            material_detail.append({
                '產品代碼': code,
                '物料代碼': mat_code,
                '單位用量': usage_per_unit,
                '產量(瓶)': qty,
                '物料成本(元)': round(total_mat_cost, 0),
            })

    logs.append({'step': '物料成本計算', 'status': '完成',
                 'detail': f'共 {len(material_detail)} 筆物料分攤'})

    # ── 彙整產品成本 ──
    results = []
    for code in PRODUCT_CODES:
        qty = prod_vol.get(code, 0)
        lc = labor_cost[code]
        mc = material_cost[code]
        total = lc + mc
        unit_cost = total / qty if qty > 0 else 0
        results.append({
            '產品代碼': code,
            '產品名稱': PRODUCT_NAMES[code],
            '本月產量(瓶)': int(qty),
            '直工薪資分攤(元)': round(lc, 0),
            '物料成本(元)': round(mc, 0),
            '製造費用合計(元)': round(total, 0),
            '單瓶成本(元)': round(unit_cost, 2),
        })

    logs.append({'step': '產品成本彙整', 'status': '完成',
                 'detail': f'共 {len(results)} 項產品'})

    # ── 輸出 Excel ──
    out_path = _write_excel(results, labor_detail, material_detail, output_dir)
    logs.append({'step': '輸出報表', 'status': '完成', 'detail': out_path})

    return {
        'success': True,
        'results': results,
        'labor_detail': labor_detail,
        'material_detail': material_detail,
        'logs': logs,
        'output_path': out_path,
    }


def _write_excel(results, labor_detail, material_detail, output_dir):
    wb = Workbook()
    t = thin_border()
    MONTH = '2026年05月'

    # ── Sheet1: 產品成本彙整表 ──
    ws1 = wb.active
    ws1.title = '產品成本彙整表'
    _set_col_widths(ws1, [14, 26, 14, 18, 16, 18, 14])

    ws1.merge_cells('A1:G1')
    c = ws1['A1']
    c.value = f'金酒公司產品成本彙整表　{MONTH}'
    c.font = Font(bold=True, size=14, color='FFFFFF')
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill(fill_type='solid', fgColor='1F4E79')

    headers = ['產品代碼', '產品名稱', '本月產量(瓶)',
               '直工薪資分攤(元)', '物料成本(元)', '製造費用合計(元)', '單瓶成本(元)']
    for j, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='BDD7EE')
        cell.border = t

    total_labor = total_mat = total_all = total_qty = 0
    for i, r in enumerate(results, 3):
        vals = [r['產品代碼'], r['產品名稱'], r['本月產量(瓶)'],
                r['直工薪資分攤(元)'], r['物料成本(元)'],
                r['製造費用合計(元)'], r['單瓶成本(元)']]
        for j, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=j, value=v)
            cell.border = t
            cell.alignment = Alignment(horizontal='center')
            if j == 3:
                cell.number_format = '#,##0'
            elif j in (4, 5, 6):
                cell.number_format = '#,##0'
            elif j == 7:
                cell.number_format = '#,##0.00'
            if i % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='EBF3FB')
        total_qty += r['本月產量(瓶)']
        total_labor += r['直工薪資分攤(元)']
        total_mat += r['物料成本(元)']
        total_all += r['製造費用合計(元)']

    tr = len(results) + 3
    ws1.cell(row=tr, column=1, value='合計').font = Font(bold=True)
    ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    for j, v in zip([3, 4, 5, 6], [total_qty, total_labor, total_mat, total_all]):
        cell = ws1.cell(row=tr, column=j, value=v)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'
        cell.fill = PatternFill(fill_type='solid', fgColor='D6E4F0')
        cell.border = t
    ws1.cell(row=tr, column=7).border = t

    # ── Sheet2: 直工薪資分攤明細 ──
    ws2 = wb.create_sheet('直工薪資分攤明細')
    _set_col_widths(ws2, [18, 14, 16, 16])
    ws2.merge_cells('A1:D1')
    c = ws2['A1']
    c.value = f'直工薪資分攤明細　{MONTH}'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill(fill_type='solid', fgColor='375623')

    for j, h in enumerate(['部門', '產品代碼', '分配工時(時)', '薪資分攤(元)'], 1):
        cell = ws2.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(fill_type='solid', fgColor='E2EFDA')
        cell.border = t

    for i, row in enumerate(labor_detail, 3):
        for j, key in enumerate(['部門', '產品代碼', '分配工時(時)', '薪資分攤(元)'], 1):
            cell = ws2.cell(row=i, column=j, value=row[key])
            cell.border = t
            cell.alignment = Alignment(horizontal='center')
            if j >= 3:
                cell.number_format = '#,##0'
            if i % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='F2F9EE')

    # ── Sheet3: 物料成本明細 ──
    ws3 = wb.create_sheet('物料成本明細')
    _set_col_widths(ws3, [14, 12, 12, 14, 16])
    ws3.merge_cells('A1:E1')
    c = ws3['A1']
    c.value = f'物料成本明細　{MONTH}'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill(fill_type='solid', fgColor='7B3F00')

    for j, h in enumerate(['產品代碼', '物料代碼', '單位用量', '產量(瓶)', '物料成本(元)'], 1):
        cell = ws3.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
        cell.border = t

    for i, row in enumerate(material_detail, 3):
        for j, key in enumerate(['產品代碼', '物料代碼', '單位用量', '產量(瓶)', '物料成本(元)'], 1):
            cell = ws3.cell(row=i, column=j, value=row[key])
            cell.border = t
            cell.alignment = Alignment(horizontal='center')
            if j in (3, 4, 5):
                cell.number_format = '#,##0.##'
            if i % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='FFF2EC')

    out_path = os.path.join(output_dir, f'產品成本表_{MONTH}.xlsx')
    wb.save(out_path)
    return out_path


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == '__main__':
    result = calculate()
    if result['success']:
        print('\n✅ 計算完成！')
        print(f'\n輸出檔案：{result["output_path"]}')
        print('\n── 產品成本彙整 ──')
        for r in result['results']:
            print(f"  {r['產品名稱']}")
            print(f"    產量：{r['本月產量(瓶)']:,} 瓶  |  單瓶成本：{r['單瓶成本(元)']:.2f} 元  |  合計：{r['製造費用合計(元)']:,.0f} 元")
        for log in result['logs']:
            print(f"  [{log['status']}] {log['step']}: {log['detail']}")
