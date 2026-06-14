"""
產生營業成本表所需的虛擬資料
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
MONTH = '2026年05月'

PRODUCTS = [
    ('KL-58-05',  '金門高粱酒 58度 0.5L',   '瓶'),
    ('KL-58-075', '金門高粱酒 58度 0.75L',  '瓶'),
    ('KL-38-05',  '金門高粱酒 38度 0.5L',   '瓶'),
    ('KL-52-06',  '金門特級高粱酒 52度 0.6L','瓶'),
]

def thin():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, cols, title, color='1F4E79'):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill(fill_type='solid', fgColor=color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def style_col_headers(ws, row, headers, color='BDD7EE'):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.fill = PatternFill(fill_type='solid', fgColor=color)
        c.border = thin()

def style_cell(cell, fmt=None):
    cell.border = thin()
    cell.alignment = Alignment(horizontal='center')
    if fmt: cell.number_format = fmt


# ── 06_期初製成品盤存表 ──────────────────────────
def gen_opening_stock():
    wb = Workbook()
    ws = wb.active
    ws.title = '期初製成品盤存表'
    for col, w in zip('ABCDEFG', [14, 24, 8, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 7, f'金酒公司期初製成品盤存表　{MONTH}　(月初)')
    style_col_headers(ws, 2, ['產品代碼','產品名稱','單位','期初數量(瓶)','單位成本(元)','期初金額(元)','備註'])

    # 期初庫存（上月結轉）
    data = [
        ('KL-58-05',  '金門高粱酒 58度 0.5L',   '瓶', 12500, 19.05),
        ('KL-58-075', '金門高粱酒 58度 0.75L',  '瓶',  8200, 26.10),
        ('KL-38-05',  '金門高粱酒 38度 0.5L',   '瓶', 15300, 16.55),
        ('KL-52-06',  '金門特級高粱酒 52度 0.6L','瓶',  5800, 23.80),
    ]
    for i, (code, name, unit, qty, unit_cost) in enumerate(data, 3):
        amount = round(qty * unit_cost, 0)
        row = [code, name, unit, qty, unit_cost, amount, '上月結轉']
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            style_cell(c)
            if j == 4: c.number_format = '#,##0'
            if j == 5: c.number_format = '#,##0.00'
            if j == 6: c.number_format = '#,##0'
        if i % 2 == 0:
            for j in range(1, 8):
                ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor='EBF3FB')

    # 合計
    tr = len(data) + 3
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    ws.cell(row=tr, column=1, value='合計').font = Font(bold=True)
    total = sum(round(d[3]*d[4], 0) for d in data)
    c = ws.cell(row=tr, column=6, value=total)
    c.font = Font(bold=True); c.number_format = '#,##0'
    c.fill = PatternFill(fill_type='solid', fgColor='D6E4F0')
    style_cell(c)

    path = os.path.join(OUTPUT_DIR, '06_期初製成品盤存表.xlsx')
    wb.save(path); print(f'✓ {path}')


# ── 07_製造費用傳票彙總 ──────────────────────────
def gen_overhead():
    wb = Workbook()
    ws = wb.active
    ws.title = '製造費用傳票彙總'
    for col, w in zip('ABCDE', [12, 24, 20, 14, 20]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 5, f'金酒公司製造費用傳票彙總　{MONTH}', color='375623')
    style_col_headers(ws, 2, ['傳票號碼','費用項目','費用說明','金額(元)','所屬廠別'], color='E2EFDA')

    data = [
        ('V2605-001', '水電費',     '金城廠5月水電費用',      285000, '金城廠'),
        ('V2605-002', '水電費',     '金寧廠5月水電費用',      312000, '金寧廠'),
        ('V2605-003', '折舊費',     '釀造設備月折舊',         480000, '全廠'),
        ('V2605-004', '折舊費',     '包裝設備月折舊',         195000, '全廠'),
        ('V2605-005', '維修費',     '設備定期保養維護',        68000, '金城廠'),
        ('V2605-006', '維修費',     '輸送管線修繕',            42000, '金寧廠'),
        ('V2605-007', '間接人工費', '廠務管理人員薪資',       320000, '全廠'),
        ('V2605-008', '消耗品',     '清潔用品、包材耗材',      35000, '全廠'),
        ('V2605-009', '保險費',     '廠房財產保險月攤',        28000, '全廠'),
        ('V2605-010', '其他製費',   '雜項製造費用',            15000, '全廠'),
    ]
    for i, row_data in enumerate(data, 3):
        for j, v in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=v)
            style_cell(c)
            if j == 4: c.number_format = '#,##0'
        if i % 2 == 0:
            for j in range(1, 6):
                ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor='F2F9EE')

    tr = len(data) + 3
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
    ws.cell(row=tr, column=1, value='合計').font = Font(bold=True)
    total = sum(d[3] for d in data)
    c = ws.cell(row=tr, column=4, value=total)
    c.font = Font(bold=True); c.number_format = '#,##0'
    c.fill = PatternFill(fill_type='solid', fgColor='D5E8C4')
    style_cell(c)

    path = os.path.join(OUTPUT_DIR, '07_製造費用傳票彙總.xlsx')
    wb.save(path); print(f'✓ {path}')


# ── 08_期末製成品盤存表 ──────────────────────────
def gen_closing_stock():
    wb = Workbook()
    ws = wb.active
    ws.title = '期末製成品盤存表'
    for col, w in zip('ABCDEFG', [14, 24, 8, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = w

    style_header(ws, 1, 7, f'金酒公司期末製成品盤存表　{MONTH}　(月末)', color='7B3F00')
    style_col_headers(ws, 2, ['產品代碼','產品名稱','單位','期末數量(瓶)','單位成本(元)','期末金額(元)','備註'], color='FCE4D6')

    data = [
        ('KL-58-05',  '金門高粱酒 58度 0.5L',   '瓶', 10800, 19.16),
        ('KL-58-075', '金門高粱酒 58度 0.75L',  '瓶',  7500, 26.28),
        ('KL-38-05',  '金門高粱酒 38度 0.5L',   '瓶', 13900, 16.68),
        ('KL-52-06',  '金門特級高粱酒 52度 0.6L','瓶',  5200, 23.93),
    ]
    for i, (code, name, unit, qty, unit_cost) in enumerate(data, 3):
        amount = round(qty * unit_cost, 0)
        row = [code, name, unit, qty, unit_cost, amount, '本月結轉']
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            style_cell(c)
            if j == 4: c.number_format = '#,##0'
            if j == 5: c.number_format = '#,##0.00'
            if j == 6: c.number_format = '#,##0'
        if i % 2 == 0:
            for j in range(1, 8):
                ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor='FFF2EC')

    tr = len(data) + 3
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    ws.cell(row=tr, column=1, value='合計').font = Font(bold=True)
    total = sum(round(d[3]*d[4], 0) for d in data)
    c = ws.cell(row=tr, column=6, value=total)
    c.font = Font(bold=True); c.number_format = '#,##0'
    c.fill = PatternFill(fill_type='solid', fgColor='F8C8A8')
    style_cell(c)

    path = os.path.join(OUTPUT_DIR, '08_期末製成品盤存表.xlsx')
    wb.save(path); print(f'✓ {path}')


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f'\n產生營業成本表虛擬資料\n')
    gen_opening_stock()
    gen_overhead()
    gen_closing_stock()
    print('\n✅ 完成！')
